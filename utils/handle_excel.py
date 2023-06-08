import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import namedtuple

from common.dir_config import CSVFILEPATH, EXCELFILEPATH


class HandleExcel(object):
    def __init__(self, filename, sheet_name=None):
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[
            self.sheet_name] if self.sheet_name is not None else self.wb.active
        self.sheet_head_tuple = tuple(
            self.ws.iter_rows(max_row=self.ws.min_row, values_only=True))[0]
        self.cases_list = []
        self.Cases = namedtuple("cases", self.sheet_head_tuple)

    def get_all_cases(self):
        for tuple_data in self.ws.iter_rows(min_row=self.ws.min_row + 1,
                                            values_only=True):
            self.cases_list.append(self.Cases(*tuple_data))
        return self.cases_list

    def get_one_case(self, row):
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            self.logger.error("Неверный номер входящей строки. Это должно быть целое число, большее 1")

    def write_file(self, row, actul_result, result_status):
        if isinstance(row, int) and (self.ws.min_row + 1 <= row <= self.ws.max_row):
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "actual") + 1, value=actul_result)
            self.ws.cell(row=row, column=self.sheet_head_tuple.index(
                "result") + 1, value=result_status)
            self.wb.save(self.filename)
        else:
            self.logger.error("Не удалось записать файл, пожалуйста, убедитесь, что в файле есть эта ячейка")


class WriteExcel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Cases"
        self.ws.append(["Case", "Title", "Description", "Method", "Host", "Port", "URL", "Params", "Response"])

        header_row = self.ws[1]
        header_font = Font(bold=True, color='FFFFFFFF', name="Bahnschrift SemiBold")
        header_fill = PatternFill(fill_type="solid", fgColor='808080')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        column_widths = [15, 90, 90, 8, 10, 8, 50, 70, 90]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col_num)
            self.ws.column_dimensions[column_letter].width = width

    def write(self, row_n, col_n, value):
        cell = self.ws.cell(row=row_n, column=col_n)
        cell.value = value

        if col_n == 12 or value in ['fail', 'error']:
            cell.font = Font(color='808080', name="Bahnschrift SemiBold", size=12, bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor='FFFFFFFF')
        elif value == 'pass':
            cell.font = Font(color="000000", name="Bahnschrift SemiBold")

    def save(self):
        self.wb.save(self.filename)

    def xlsx_to_csv_pd(self):
        data_xls = pd.read_excel(self.filename, index_col=0)
        data_xls.to_csv(CSVFILEPATH, encoding='utf-8_sig')


w = WriteExcel(EXCELFILEPATH)
